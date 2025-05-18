# Pipeline ETL: Cotação do Dólar
# 1. Extrair dados da cotação do dólar dos últimos dias via API.
# 2. Transformar: Selecionar colunas relevantes, formatar datas e valores.
# 3. Carregar: Salvar em um arquivo XSLX local e atualizar um gráfico de análise do comportamento do dolar.
# Para desenvolver essa pipeline ETL foi necessário a utilização da API (https://docs.awesomeapi.com.br/api-de-moedas)

import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Caminho do arquivo Excel e nome da aba onde os dados serão atualizados
arquivo_excel = "Histórico Dolar.xlsx"
aba_dados = "Gráficos de análise"

# --------- ETAPA 1: EXTRAÇÃO ---------
# Faz a requisição para a API pública da AwesomeAPI com os dados dos últimos 7 dias do dólar
url = "https://economia.awesomeapi.com.br/json/daily/USD-BRL/7"
response = requests.get(url)

# Verifica se a requisição foi bem-sucedida
if response.status_code == 200:
    dados = response.json()

    # --------- ETAPA 2: TRANSFORMAÇÃO ---------
    # Cria um DataFrame com apenas as colunas necessárias
    df = pd.DataFrame(dados)[["timestamp", "bid", "ask", "high", "low"]].copy()

    # Converte a coluna timestamp (em segundos) para data legível
    df["data"] = df["timestamp"].apply(lambda ts: datetime.fromtimestamp(int(ts)).date())

    # Reorganiza a ordem das colunas
    df = df[["data", "bid", "ask", "high", "low"]]

    # Converte os valores de string para float (formato numérico contábil)
    for col in ["bid", "ask", "high", "low"]:
        df[col] = df[col].astype(float)

    # Ordena o DataFrame por data, do mais antigo para o mais recente
    df = df.sort_values("data")

    # --------- ETAPA 3: CARGA (LOAD) ---------
    # Abre a planilha existente, preservando objetos como gráficos
    wb = load_workbook(arquivo_excel)
    ws = wb[aba_dados]

    # Determina o intervalo de dados antigos a serem limpos (sem apagar cabeçalho ou gráfico)
    max_row = ws.max_row
    max_col = ws.max_column

    # Limpa os dados antigos da planilha (mantendo cabeçalho e gráfico)
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.value = None

    # Insere os novos dados na planilha a partir da linha 2
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Salva a planilha atualizada com os novos dados
    wb.save(arquivo_excel)

    print("Dados atualizados com sucesso!")

else:
    print(f"Erro ao acessar API: {response.status_code}")